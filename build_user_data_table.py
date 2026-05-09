"""
User data table for veteran events platform (SKELAR Hackathon brief).
Design principles:
- Minimum required fields at signup (no bureaucratic wall)
- Progressive enrichment afterward
- Explicit mapping to event-matching logic
- State hand-off friendly: simple types, no exotic structures
- Inclusivity: fields capture all 4 persona barrier types
- Privacy: minimum PII, clear consent
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = Workbook()

# ------- STYLES -------
HEADER_FILL = PatternFill("solid", start_color="1F3A5F")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", start_color="E8EEF7")
SECTION_FONT = Font(name="Arial", bold=True, color="1F3A5F", size=11)
BODY_FONT = Font(name="Arial", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
REQUIRED_FILL = PatternFill("solid", start_color="FFF2CC")  # yellow tint for required
OPTIONAL_FILL = PatternFill("solid", start_color="FFFFFF")

def style_header_row(ws, row, last_col):
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

def style_body(ws, start_row, end_row, last_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, last_col + 1):
            cell = ws.cell(row=r, column=c)
            if not cell.font or cell.font.name != "Arial":
                cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = BORDER

# ============================================================
# SHEET 1: SCHEMA
# ============================================================
ws = wb.active
ws.title = "1. Schema"

ws["A1"] = "User Data Schema — Veteran Events Platform"
ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1F3A5F")
ws.merge_cells("A1:H1")

ws["A2"] = ("Field design driven by the 4 personas in the brief (Дмитро, Катерина, Михайло, Василь). "
            "Required fields kept to a minimum to avoid bureaucratic onboarding. "
            "Optional fields populate progressively and feed the event-matching engine.")
ws["A2"].font = Font(name="Arial", italic=True, size=10, color="555555")
ws["A2"].alignment = WRAP
ws.merge_cells("A2:H2")
ws.row_dimensions[2].height = 32

headers = ["Field", "Group", "Type", "Required at signup",
           "Source / how collected", "Example value", "Why it matters (event matching)", "Persona it serves"]
for i, h in enumerate(headers, 1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, len(headers))
ws.row_dimensions[4].height = 38

schema_rows = [
    # === IDENTITY & VERIFICATION ===
    ("user_id", "Identity", "UUID", "Yes (auto)", "System-generated on signup", "u_8f3c2a1b...", "Stable join key for events, attendance, saved items.", "All"),
    ("display_name", "Identity", "Text (≤40)", "No", "Self-entered. Default = first name only.", "Дмитро", "Personal greeting only. Never used in matching. Privacy default = first name.", "All"),
    ("year_of_birth", "Identity", "Integer", "No", "Self-entered (year, not full DOB)", "1992", "Drives age-appropriate event filtering (e.g., 46+ programs differ in tone). Year only = less PII.", "All"),
    ("gender", "Identity", "Enum: female / male / not_specified", "No", "Self-entered, never required", "female", "Surfaces women-veteran communities — Катерина's invisibility bar.", "Катерина"),
    ("verification_status", "Verification", "Enum: verified / pending / self_declared", "Yes", "Set by verification flow", "self_declared", "Some events restrict to verified only; default access stays low-friction.", "All"),
    ("verification_method", "Verification", "Enum: diia / photo_id / qr_code / self_declaration / partner_org", "Yes", "Chosen by user at verification step", "self_declaration", "Lets the platform start with self-declaration (MVP) and upgrade to Дія later without schema changes.", "All"),
    ("verification_date", "Verification", "Date", "No", "Stamped on successful verification", "2026-04-12", "Auditable for partner orgs that require recent verification.", "All"),

    # === SERVICE CONTEXT ===
    ("service_type", "Service", "Enum: contract / mobilized / volunteer_formation / national_guard / medic / other", "No", "Self-entered (single choice)", "contract", "Some events target specific service contexts. Optional to avoid gating.", "Дмитро, Катерина"),
    ("service_status", "Service", "Enum: serving / demobilized / discharged_medical / reserve", "No", "Self-entered", "demobilized", "Time-since-return drives the type of activity surfaced (just-returned vs settled).", "Дмитро, Михайло"),
    ("return_year", "Service", "Integer (year)", "No", "Self-entered", "2025", "First-3-months users see a different home screen than 2-years-out users.", "Дмитро"),

    # === LOCATION ===
    ("region", "Location", "Enum: 24 oblasts + Kyiv", "Yes", "Self-entered (dropdown) or geo-suggested", "Poltavska", "Primary geographic filter. Critical for Василь's 'this is all for Київ' bar.", "Василь, Михайло"),
    ("city", "Location", "Text", "No", "Self-entered or geo-suggested", "Hlobyne", "Exact city for proximity-based event sorting.", "Василь"),
    ("settlement_size", "Location", "Enum: capital / regional_centre / district_centre / village", "No", "Auto-derived from city", "district_centre", "Lets matching engine prioritise online-first events for small-settlement users.", "Василь"),
    ("max_travel_distance_km", "Location", "Integer (default 25)", "No", "Profile setting (slider)", "30", "Filters events by realistic reach. Honours transport limits without asking about them directly.", "Василь, Михайло"),

    # === ACCESSIBILITY ===
    ("has_accessibility_needs", "Accessibility", "Boolean", "No", "Self-entered toggle", "TRUE", "Master switch that unlocks the accessibility filters below.", "Михайло"),
    ("accessibility_needs", "Accessibility", "Multi-select: mobility / vision / hearing / cognitive / sensory_overload / none", "No", "Self-entered (multi-checkbox)", "mobility", "Filters out venues without ramps, surfaces adaptive sport, etc. Михайло's main bar.", "Михайло"),
    ("mobility_aid", "Accessibility", "Enum: none / prosthetic / wheelchair / cane / crutches / other", "No", "Self-entered", "prosthetic", "Distinguishes 'adaptive' vs 'fully accessible' venue requirements.", "Михайло"),

    # === PREFERENCES (drive AI matching) ===
    ("preferred_event_types", "Preferences", "Multi-select: sport / outdoor / art_culture / education / community / family / volunteering / mental_wellbeing / craft_workshop", "No", "Onboarding quiz (tap-to-select cards)", "sport, community", "Core matching signal. Captured visually, not via long forms — reduces friction for low digital literacy.", "All"),
    ("preferred_format", "Preferences", "Multi-select: offline / online / hybrid", "No", "Onboarding toggle", "offline, online", "Online-first matters for Василь and small-settlement users.", "Василь"),
    ("preferred_intensity", "Preferences", "Enum: calm / moderate / active / adaptive", "No", "Visual scale 1–4", "moderate", "Distinguishes 'quiet meet-up' from 'high-intensity sport'. Avoids surfacing a 5K run to someone seeking quiet.", "All"),
    ("preferred_company", "Preferences", "Enum: only_veterans / mixed / women_only / family_friendly / no_preference", "No", "Onboarding question", "only_veterans", "Direct answer to Дмитро's 'people who don't ask how it was there' and Катерина's women-only need.", "Дмитро, Катерина"),
    ("family_can_join", "Preferences", "Boolean", "No", "Profile toggle", "TRUE", "Filters family-friendly events when the partner is the one nudging participation (Михайло scenario).", "Михайло"),

    # === ENGAGEMENT (auto-tracked) ===
    ("preferred_channel", "Engagement", "Enum: telegram / email / sms / push", "Yes (default = telegram)", "Auto-set; user can change", "telegram", "Drives reminders and digest delivery — Telegram is the lowest-friction channel for Ukrainian users.", "All"),
    ("digital_literacy_signal", "Engagement", "Enum: low / medium / high", "No", "Auto-inferred from completed-action ratio (or manual)", "low", "Switches UI to simplified mode (large text, fewer steps). Never shown to user as a label.", "Василь"),
    ("last_active_at", "Engagement", "Timestamp", "Auto", "System", "2026-05-08T19:22:00Z", "Re-engagement triggers; stale profiles get a check-in nudge instead of more event blasts.", "All"),
    ("attended_events_count", "Engagement", "Integer", "Auto", "Incremented on attendance confirmation", "3", "Used to escalate from solo-friendly events → community events as confidence grows.", "All"),
    ("saved_events_count", "Engagement", "Integer", "Auto", "Incremented on save", "5", "Signal for 'interested but not committing' — triggers gentle nudge or smaller-format alternatives.", "Дмитро"),

    # === CONSENT & PRIVACY ===
    ("consent_data_processing", "Consent", "Boolean", "Yes", "Required at signup", "TRUE", "GDPR/Ukrainian PDP compliance. Hard requirement.", "All"),
    ("consent_marketing", "Consent", "Boolean", "No (default FALSE)", "Opt-in checkbox", "FALSE", "Separates transactional reminders from promotional outreach. Default off.", "All"),
    ("profile_visibility", "Consent", "Enum: private / first_name_only / full_profile", "No (default first_name_only)", "Profile setting", "first_name_only", "Privacy default avoids exposing veterans to other users by accident.", "Катерина"),
]

start = 5
for i, row in enumerate(schema_rows):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=start + i, column=c, value=val)
        cell.alignment = WRAP
        cell.font = BODY_FONT
        cell.border = BORDER
        # Tint required-at-signup rows
        if row[3].lower().startswith("yes"):
            cell.fill = REQUIRED_FILL

# Column widths
widths = {"A": 26, "B": 14, "C": 28, "D": 18, "E": 30, "F": 22, "G": 50, "H": 18}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Row heights — let Excel auto-fit usually works, but fix the schema rows generously
for r in range(start, start + len(schema_rows)):
    ws.row_dimensions[r].height = 48

# Legend
legend_row = start + len(schema_rows) + 2
ws.cell(row=legend_row, column=1, value="Legend:").font = Font(name="Arial", bold=True, size=10)
ws.cell(row=legend_row, column=2, value="Required-at-signup fields").fill = REQUIRED_FILL
ws.cell(row=legend_row, column=2).font = BODY_FONT
ws.cell(row=legend_row, column=2).border = BORDER

# ============================================================
# SHEET 2: SAMPLE USERS
# ============================================================
ws2 = wb.create_sheet("2. Sample users")

ws2["A1"] = "Sample user records — covering all 4 personas + edge cases"
ws2["A1"].font = Font(name="Arial", bold=True, size=14, color="1F3A5F")
ws2.merge_cells("A1:Q1")

ws2["A2"] = ("Each row populates the schema with realistic values. Notice how fields combine to expose / hide events: "
             "Михайло's mobility need + small_travel_distance trims the candidate set; Катерина's gender + women_only "
             "preference unlocks a different community. AI matching reads the same columns.")
ws2["A2"].font = Font(name="Arial", italic=True, size=10, color="555555")
ws2["A2"].alignment = WRAP
ws2.merge_cells("A2:Q2")
ws2.row_dimensions[2].height = 38

user_headers = [
    "user_id", "display_name", "year_of_birth", "gender",
    "verification_status", "service_type", "service_status", "return_year",
    "region", "settlement_size", "max_travel_km",
    "accessibility_needs", "preferred_event_types", "preferred_intensity",
    "preferred_company", "preferred_channel", "Persona / scenario note"
]
for i, h in enumerate(user_headers, 1):
    ws2.cell(row=4, column=i, value=h)
style_header_row(ws2, 4, len(user_headers))
ws2.row_dimensions[4].height = 36

users = [
    # === The 4 brief personas ===
    ("u_dmytro_001", "Дмитро", 1992, "male", "self_declared", "contract", "demobilized", 2025,
     "Dnipropetrovska", "regional_centre", 25, "none", "sport, community", "moderate",
     "only_veterans", "telegram",
     "Persona: Дмитро. Just-returned, psychological+info bar. Surfaces: low-intensity sport meets, veteran-only fitness."),
    ("u_kateryna_002", "Катерина", 1998, "female", "verified", "medic", "demobilized", 2024,
     "Lvivska", "capital", 30, "none", "art_culture, community, mental_wellbeing", "calm",
     "women_only", "telegram",
     "Persona: Катерина. Identity bar. Surfaces: women-veteran circles, art workshops, no over-heroising tone."),
    ("u_mykhailo_003", "Михайло", 1979, "male", "verified", "mobilized", "discharged_medical", 2023,
     "Kharkivska", "regional_centre", 15, "mobility", "sport, family", "adaptive",
     "family_friendly", "telegram",
     "Persona: Михайло. Physical+info bar. Surfaces: adaptive sport, family-friendly events with venue accessibility verified."),
    ("u_vasyl_004", "Василь", 1973, "male", "self_declared", "mobilized", "demobilized", 2025,
     "Poltavska", "district_centre", 40, "none", "outdoor, community", "moderate",
     "mixed", "sms",
     "Persona: Василь. Geographic+digital bar. SMS channel + larger travel radius + simplified UI auto-applied."),

    # === Edge cases ===
    ("u_olena_005", "Олена", 1986, "female", "verified", "national_guard", "serving", None,
     "Kyivska", "capital", 20, "none", "education, mental_wellbeing", "calm",
     "women_only", "telegram",
     "Active service member exploring future activities. Sees online-first events."),
    ("u_andriy_006", "Андрій", 1995, "male", "self_declared", "volunteer_formation", "demobilized", 2024,
     "Odeska", "regional_centre", 30, "hearing", "outdoor, craft_workshop", "active",
     "mixed", "telegram",
     "Hearing accessibility need — filters out events without sign-language or captions."),
    ("u_iryna_007", "Ірина", 1990, "female", "verified", "medic", "demobilized", 2026,
     "Vinnytska", "regional_centre", 25, "none", "family, community", "calm",
     "family_friendly", "email",
     "Just returned, family-first preference. Email is her preferred channel."),
    ("u_petro_008", "Петро", 1968, "male", "self_declared", "mobilized", "demobilized", 2025,
     "Zakarpatska", "village", 50, "cognitive", "outdoor, craft_workshop", "calm",
     "mixed", "sms",
     "Older, village resident, cognitive accessibility. Auto-simplified UI; SMS reminders."),
    ("u_yulia_009", "Юлія", 2000, "female", "verified", "contract", "serving", None,
     "Lvivska", "capital", 15, "none", "art_culture, education", "moderate",
     "women_only", "telegram",
     "Young woman, still serving. Art + education community focus."),
    ("u_serhiy_010", "Сергій", 1982, "male", "verified", "mobilized", "discharged_medical", 2024,
     "Donetska", "district_centre", 60, "mobility, vision", "education, community", "calm",
     "only_veterans", "telegram",
     "Multiple accessibility needs + frontline region. Larger travel radius — limited local options."),
    ("u_natalia_011", "Наталія", 1995, "female", "verified", "national_guard", "demobilized", 2025,
     "Cherkaska", "district_centre", 35, "sensory_overload", "craft_workshop, mental_wellbeing", "calm",
     "women_only", "telegram",
     "Sensory overload — filters out loud / large-crowd events. Quiet workshops surfaced."),
    ("u_oleh_012", "Олег", 1990, "male", "self_declared", "mobilized", "demobilized", 2026,
     "Zaporizka", "regional_centre", 25, "none", "sport, volunteering", "active",
     "only_veterans", "telegram",
     "Ready-to-give-back profile. Volunteering events surfaced alongside sport."),
]

for r, u in enumerate(users):
    for c, val in enumerate(u, 1):
        cell = ws2.cell(row=5 + r, column=c, value=val)
        cell.alignment = WRAP
        cell.font = BODY_FONT
        cell.border = BORDER

ws2.row_dimensions[1].height = 22
for r in range(5, 5 + len(users)):
    ws2.row_dimensions[r].height = 60

user_widths = [16, 14, 11, 10, 14, 18, 18, 11, 17, 15, 12, 22, 28, 14, 16, 14, 50]
for i, w in enumerate(user_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# Freeze header
ws2.freeze_panes = "A5"

# ============================================================
# SHEET 3: EVENT MATCHING LOGIC
# ============================================================
ws3 = wb.create_sheet("3. Event matching")

ws3["A1"] = "How user fields drive event matching"
ws3["A1"].font = Font(name="Arial", bold=True, size=14, color="1F3A5F")
ws3.merge_cells("A1:E1")

ws3["A2"] = ("Each user field serves a specific role in the matching pipeline: hard filter (excludes events), "
             "soft signal (re-ranks), or contextual modifier (changes copy/UI). Designed so a non-tech operator "
             "can manage events via simple tags without writing code.")
ws3["A2"].font = Font(name="Arial", italic=True, size=10, color="555555")
ws3["A2"].alignment = WRAP
ws3.merge_cells("A2:E2")
ws3.row_dimensions[2].height = 36

match_headers = ["User field", "Role in matching", "Event-side counterpart (tag)", "Example", "Hand-off note (no-code maintenance)"]
for i, h in enumerate(match_headers, 1):
    ws3.cell(row=4, column=i, value=h)
style_header_row(ws3, 4, len(match_headers))
ws3.row_dimensions[4].height = 36

match_rows = [
    ("region + max_travel_km", "Hard filter", "event.region + event.city + event.format",
     "Василь in Hlobyne with 40km radius sees Poltava events + all 'online' events.",
     "Operator just tags event with one oblast + city + 'online/offline'."),
    ("accessibility_needs", "Hard filter", "event.venue_accessibility (multi-tag)",
     "Михайло (mobility) excludes any event without 'venue_accessible' tag. Adaptive sport always shown.",
     "Operator ticks accessibility checkboxes when adding event in admin."),
    ("preferred_event_types", "Soft re-rank", "event.category (single)",
     "Sport-tagged events float to top for Дмитро.",
     "Single category dropdown per event — minimal admin overhead."),
    ("preferred_intensity", "Soft re-rank", "event.intensity",
     "'Calm' events rank higher for Катерина; 'active' events for Олег.",
     "Optional dropdown; defaults to 'moderate' if operator skips."),
    ("preferred_company", "Hard filter (only_veterans, women_only)", "event.audience",
     "Women-only events hidden from male users; veteran-only events hidden from non-verified.",
     "Two boolean flags per event: 'veterans_only', 'women_only'."),
    ("verification_status", "Hard filter (when event requires)", "event.verification_required",
     "Self-declared users still see most events; only some (e.g., subsidised travel) require verified.",
     "One checkbox per event."),
    ("digital_literacy_signal", "UI modifier", "—",
     "Auto-switches to large-text, single-column, no-jargon copy. Никогда не показується юзеру.",
     "No operator action — derived from user behaviour or self-set."),
    ("return_year", "Soft re-rank", "event.suitable_for_recently_returned",
     "Just-returned users see lower-pressure events first.",
     "One boolean tag per event."),
    ("family_can_join", "Soft signal", "event.family_friendly",
     "Family-friendly events surface for users who toggled it on.",
     "One boolean."),
    ("preferred_channel", "Routing only", "—",
     "Reminder sent via Telegram / SMS / email per user setting. No event-side change.",
     "No operator action."),
    ("attended_events_count", "Personalisation", "—",
     "0–1 attended → solo-friendly events; 5+ → community / volunteering surfaced.",
     "Auto-tracked. Operator never touches."),
]

for r, row in enumerate(match_rows):
    for c, val in enumerate(row, 1):
        cell = ws3.cell(row=5 + r, column=c, value=val)
        cell.alignment = WRAP
        cell.font = BODY_FONT
        cell.border = BORDER
    ws3.row_dimensions[5 + r].height = 56

m_widths = [24, 22, 30, 38, 42]
for i, w in enumerate(m_widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# SHEET 4: DESIGN NOTES
# ============================================================
ws4 = wb.create_sheet("4. Design notes")

ws4["A1"] = "Design rationale & hand-off considerations"
ws4["A1"].font = Font(name="Arial", bold=True, size=14, color="1F3A5F")
ws4.merge_cells("A1:B1")

notes = [
    ("Required fields kept to 5",
     "Only user_id (auto), verification_status, region, preferred_channel, consent_data_processing are mandatory at signup. "
     "Everything else fills progressively. This addresses the brief's demand for 'simple entry, no extra steps' and avoids the "
     "bureaucratic-form bar described for Василь."),
    ("Year-of-birth, not full DOB",
     "Reduces PII surface for state hand-off. Year is sufficient for age-band filtering (46+ programs) but does not store full birth date."),
    ("Verification by self-declaration first",
     "Per brief Q&A: Дія integration is a future step. Schema supports self_declaration → photo_id → diia upgrade path without changes."),
    ("Accessibility split into 3 fields",
     "has_accessibility_needs (master toggle) → accessibility_needs (multi-select) → mobility_aid (specific). "
     "Lets users disclose only as much as they want, avoiding the 'pity-the-disabled-veteran' patronising tone."),
    ("preferred_company covers identity scenarios",
     "women_only and only_veterans options directly address Катерина's invisibility and Дмитро's 'people who don't ask' need — "
     "without making the user articulate it themselves."),
    ("digital_literacy_signal is auto-derived",
     "Never shown as a label. Triggers a simplified UI under the hood. Avoids labelling users as 'low-skill' — preserves dignity."),
    ("All multi-selects are tag-style",
     "Operator side: events tagged with single dropdowns + boolean flags. No SQL, no scripting. Maintainable by one person on Google Sheets / Airtable."),
    ("Telegram-first communication",
     "Default channel = telegram. Lowest infra cost, most familiar UX for Ukrainian veterans. SMS fallback for older / village users (Василь)."),
    ("Privacy default = first_name_only",
     "Profile visibility starts at first-name-only — explicit opt-in to expose more. Protects users who don't want to be discoverable."),
    ("AI integration points",
     "1) Onboarding quiz: AI converts free-text answers ('I want quiet stuff with people like me') into preferred_event_types + "
     "preferred_intensity + preferred_company. 2) Event copy translation: AI rewrites event descriptions in plain language per user's "
     "digital_literacy_signal. 3) Re-engagement: AI generates a personal nudge based on saved_events_count + last_active_at."),
    ("State hand-off readiness",
     "All fields are flat columns, all enums are documented strings — no nested JSON, no dynamic schemas. Maps 1:1 to a Postgres "
     "table or even a Google Sheet. Non-tech operator can manage user profiles and events without a developer."),
]

for r, (title, body) in enumerate(notes):
    tcell = ws4.cell(row=3 + r * 2, column=1, value=title)
    tcell.font = Font(name="Arial", bold=True, size=11, color="1F3A5F")
    tcell.alignment = WRAP
    tcell.fill = SECTION_FILL
    bcell = ws4.cell(row=3 + r * 2, column=2, value=body)
    bcell.font = BODY_FONT
    bcell.alignment = WRAP
    ws4.row_dimensions[3 + r * 2].height = 60

ws4.column_dimensions["A"].width = 32
ws4.column_dimensions["B"].width = 95

# Save
output_path = "/sessions/awesome-amazing-archimedes/mnt/bemguide-solution/veteran_user_data_table.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
